import json
from pathlib import Path
from datetime import datetime

from django.conf import settings
from django.utils import timezone
from django.http import HttpRequest

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from robots.models import Robot


def validate_new_robot_request(request: HttpRequest) -> dict[str, str] | None:
    """Return valid JSON from `request.body` as `dict`.
    If JSON doesn't meet the requirements, raise either `TypeError` or `ValueError` with corresponding message.
    """

    # Verify that JSON is valid and uses UTF-8 encoding
    try:
        params = json.loads(request.body.decode("utf-8"))
    except UnicodeError:
        raise ValueError("Encoding must be 'utf-8'")
    except ValueError:
        raise ValueError("Invalid JSON")
    else:
        if not isinstance(params, dict):
            raise ValueError("Invalid JSON")

    # Verify that JSON has all required params as strings
    required_params = ("model", "version", "created")
    if len(params) != len(required_params):
        raise ValueError(f"Request must contain exactly {len(required_params)} params")

    for param in required_params:
        if params.get(param) is None:
            raise TypeError(f"'{param}' is missing")
        if not isinstance(params[param], str):
            raise TypeError(f"'{param}' must be a string")

    # Verify that `model` and `version` are valid
    valid_length = 2
    for param in ("model", "version"):
        if len(params[param]) != valid_length or len(params[param].strip()) != valid_length:
            raise ValueError(f"'{param}' must contain exactly {valid_length} non-whitespace characters")

    # Verify that the timestamp is in the correct format
    timestamp_format = "%Y-%m-%d %H:%M:%S"
    try:
        timestamp = datetime.strptime(params["created"], timestamp_format)
    except ValueError:
        raise ValueError(f"'created' must match the following pattern: '{timestamp_format}'")

    # Make sure there's no robot already assembled at that second
    if Robot.objects.is_assembled_at_second(timestamp):
        raise ValueError("A robot assembled at this second already exists")

    return params


def get_last_week_stats() -> dict:
    """Return summary of robot production totals for the last week"""
    data = {}
    for model in Robot.objects.models():
        data |= Robot.objects.last_week_production_summary(model=model)

    return data


def create_xlsx_file() -> Path:
    """Save data received from `get_last_week_stats()` as `.xlsx` file. Return its path."""
    wb = Workbook()
    timestamp = timezone.now()  # Will be added to filename to make it unique and informative
    data = get_last_week_stats()

    # Optimization. Create all of these only once since header is the same on every sheet.
    header = (("A1", "Модель"), ("B1", "Версия"), ("C1", "Количество за неделю"))
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    narrow_column_width, wide_column_width = 10, 25

    for model, model_data in data.items():
        # Create sheet and resize table
        ws = wb.create_sheet(model)
        ws.column_dimensions["A"].width = narrow_column_width
        ws.column_dimensions["B"].width = narrow_column_width
        ws.column_dimensions["C"].width = wide_column_width

        # Fill table header
        for cell, value in header:
            ws[cell] = value
            ws[cell].font = header_font
            ws[cell].alignment = header_alignment

        # Fill table rows
        for row_idx, version_data in enumerate(model_data, start=2):
            ws[f"A{row_idx}"], ws[f"B{row_idx}"], ws[f"C{row_idx}"] = (
                model,
                version_data["version"],
                version_data["count"],
            )

    wb.remove(wb["Sheet"])  # Remove default empty sheet from workbook

    output_folder = Path(settings.BASE_DIR, "reports")
    output_folder.mkdir(parents=True, exist_ok=True)
    output_file = output_folder / f"report_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)

    return output_file
