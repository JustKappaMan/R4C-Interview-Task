from pathlib import Path

from django.conf import settings
from django.utils import timezone as tz

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from robots.models import Robot


def _get_last_week_stats() -> dict:
    """Return summary of robot production totals for the last week"""
    data = {}
    for model in Robot.objects.models():
        data |= Robot.objects.last_week_production_summary(model=model)

    return data


def create_xlsx_file() -> Path:
    """Save data received from `_get_last_week_stats()` as `.xlsx` file. Return its path."""
    wb = Workbook()
    timestamp = tz.now()  # Will be added to filename to make it unique and informative
    data = _get_last_week_stats()

    # Optimization. Create all of these only once since header is the same on every sheet.
    header = (("A1", "Модель"), ("B1", "Версия"), ("C1", "Количество за неделю"))
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    narrow_column_width, wide_column_width = 10, 25

    for model, model_data in data.items():
        # Create sheet and resize table
        ws = wb.create_sheet(model)
        ws.column_dimensions["A"].width = narrow_column_width
        ws.column_dimensions["B"].width = narrow_column_width
        ws.column_dimensions["C"].width = wide_column_width

        # Fill table header
        for cell, value in header:
            ws[cell] = value
            ws[cell].font = header_font
            ws[cell].alignment = header_alignment

        # Fill table rows
        for row_idx, version_data in enumerate(model_data, start=2):
            ws[f"A{row_idx}"], ws[f"B{row_idx}"], ws[f"C{row_idx}"] = (
                model,
                version_data["version"],
                version_data["count"],
            )

    if len(wb.worksheets) > 1:
        wb.remove(wb["Sheet"])  # Remove default empty sheet from workbook

    output_folder = Path(settings.BASE_DIR, "reports")
    output_folder.mkdir(parents=True, exist_ok=True)
    output_file = output_folder / f"report_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)

    return output_file
